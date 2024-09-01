from openpyxl import load_workbook # type: ignore
from yattag import Doc, indent

# Loading our Excell file
wb = load_workbook("demo_database.xlsx")

# Creating the sheet 1 object
ws = wb.worksheets[0]

# Returning returns a triplet
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version= "1.0" encoding= "UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs= "http://www.w3.org/2001/XMLSchema"></xs:schema>'

# Append the string document
doc.asis(xml_header)
doc.asis(xml_schema)

min_row, max_row = int(input("Enter minimum row: ")), int(input("and maximum row number: "))
min_col, max_col = int(input("Enter minimum column: ")), int(input("and maximum column number: "))

# iterating rows for getting the values of each row
with tag('People'):
    for row in ws.iter_rows(min_row, max_row, min_col, max_col):
        row = ([cell.value for cell in row])
        with tag('Person'):
            with tag('First_Name'):
                text(row[0])
            with tag('Last_Name'):
                text(row[1])
            with tag('Gender'):
                text(row[2])
            with tag('Country'):
                text(row[3])
            with tag('Age'):
                text(row[4])
            with tag('Date'):
                text(row[5])

result = indent(
    doc.getvalue(), 
    indentation="   ",
    indent_text=True
)

with open("Output.xml", "w") as f:
    f.write(result)
    